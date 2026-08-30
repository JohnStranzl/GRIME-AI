#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Author: John Edward Stranzl, Jr.
# Affiliation(s): University of Nebraska-Lincoln, Blade Vision Systems, LLC
# Contact: jstranzl2@huskers.unl.edu, johnstranzl@gmail.com
# License: Apache License, Version 2.0, http://www.apache.org/licenses/LICENSE-2.0

"""
Analyze the relationship between segmented ROI water area and co-located
NWIS sensor parameters, and write a multi-tab Excel report.

Input: the ROI feature-extraction DataFrame with correlated sensor columns
(one column per parameter, named "<description> [<ts_id>_<param>]").

Output workbook:
  - one tab per sensor parameter: the paired data plus a scatter chart of
    the sensor value (y) against ROI Area Percentage (x)
  - an Assessment tab: correlation statistics for every parameter and a
    rule-based narrative interpreting what the numbers suggest at the site

The narrative is heuristic. It states what the statistics are consistent
with; it does not claim to establish physical causation.
"""

import re
import numpy as np
import pandas as pd

PARAM_COL_RE = re.compile(r"\[(\d+)_(\d{5})\]$")

# NWIS parameter codes with known physical relationships to water surface area
# Analysis technique declared per parameter class (agreed methodology):
#   00065 stage:     raw-value regression is authoritative; sensitivity reported
#                    in %-area per unit AND per stage percentile.
#   00060 discharge: rank statistics authoritative (stage-derived through a
#                    nonlinear rating); raw Pearson understates by construction.
#   everything else: rank statistics, interpret cautiously (pending discussion).
METHODS = {
    "00065": "Raw-value regression (linear vs power by R^2); sensitivity in %/unit and %/percentile",
    "00060": "Rank-based (Spearman authoritative); stage-derived - interpret via stage",
}
METHOD_DEFAULT = "Rank statistics (interpret cautiously - technique not yet finalized)"

CODE_HINTS = {
    "00065": "gage height",
    "00060": "discharge",
    "00010": "water temperature",
    "00020": "air temperature",
    "00095": "specific conductance",
    "00300": "dissolved oxygen",
    "63680": "turbidity",
}


class SensorROIAnalysis:

    def __init__(self):
        self.className = "SensorROIAnalysis"

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def write_report(self, df: pd.DataFrame, xlsx_path: str,
                     roi_col: str = "ROI Area Percentage") -> str:
        """Write the sensor-vs-ROI analysis workbook. Returns xlsx_path."""
        if roi_col not in df.columns:
            candidates = [c for c in df.columns if "Area Percentage" in str(c)]
            if not candidates:
                raise ValueError(f"Column '{roi_col}' not found in the feature DataFrame.")
            roi_col = candidates[0]

        param_cols = [c for c in df.columns if PARAM_COL_RE.search(str(c))]
        if not param_cols:
            raise ValueError("No sensor parameter columns found "
                             "(expected names ending in [tsid_code]).")

        work = df.copy()
        work["_roi"] = pd.to_numeric(work[roi_col], errors="coerce")
        work["_ts"] = self._timestamps(work)
        work = work.sort_values("_ts", kind="stable").reset_index(drop=True)

        all_stats = []
        for col in param_cols:
            values = pd.to_numeric(work[col], errors="coerce")
            all_stats.append(self._param_stats(col, values, work["_roi"]))

        # If both discharge and stage are present, state whether discharge adds
        # independent information beyond stage (it rarely does: 00060 is
        # rating-derived from 00065 at most gages).
        by_code = {st["code"]: st for st in all_stats}
        if "00060" in by_code and "00065" in by_code:
            q, h = by_code["00060"], by_code["00065"]
            if not np.isnan(q["spearman"]) and not np.isnan(h["spearman"]):
                if abs(q["spearman"] - h["spearman"]) < 0.05:
                    q["assessment"] += (f" Spearman vs discharge ({q['spearman']:.2f}) and vs "
                                        f"gage height ({h['spearman']:.2f}) are nearly identical: "
                                        f"discharge adds no independent information beyond stage "
                                        f"at this site.")

        self._write_workbook(xlsx_path, work, roi_col, param_cols, all_stats)
        return xlsx_path

    # ==================================================================================================================
    #
    # ==================================================================================================================
    @staticmethod
    def _timestamps(df: pd.DataFrame) -> pd.Series:
        """Best-effort chronological timestamps for limb (rising/falling) analysis."""
        if "Capture Date" in df.columns and "Capture Time" in df.columns:
            ts = pd.to_datetime(df["Capture Date"].astype(str) + " " + df["Capture Time"].astype(str),
                                errors="coerce")
            if ts.notna().any():
                return ts
        if "Sensor Time (UTC)" in df.columns:
            ts = pd.to_datetime(df["Sensor Time (UTC)"], errors="coerce")
            if ts.notna().any():
                return ts
        return pd.Series(pd.RangeIndex(len(df)), index=df.index)

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _param_stats(self, col: str, values: pd.Series, roi: pd.Series) -> dict:
        """Correlation statistics for one parameter against ROI area."""
        m = PARAM_COL_RE.search(str(col))
        code = m.group(2) if m else ""
        mask = values.notna() & roi.notna()
        v, a = values[mask], roi[mask]
        n = int(mask.sum())

        stats = {
            "column": col, "code": code, "n": n,
            "pearson": np.nan, "spearman": np.nan,
            "log_pearson": np.nan, "loglog_slope": np.nan, "loglog_r2": np.nan,
            "pearson_rising": np.nan, "pearson_falling": np.nan,
            "pearson_low_half": np.nan, "pearson_high_half": np.nan,
            "linear_slope": np.nan, "linear_r2": np.nan, "pctile_slope": np.nan,
        }
        if n < 3 or v.nunique() < 3 or a.nunique() < 3:
            stats["assessment"] = (f"Insufficient paired data (N={n}) for a meaningful "
                                   f"assessment of this parameter.")
            return stats

        stats["pearson"]  = float(a.corr(v, method="pearson"))
        stats["spearman"] = float(a.corr(v, method="spearman"))

        # Linear fit: area = m*value + c  (m = sensitivity in %-area per unit)
        try:
            m_lin, c_lin = np.polyfit(v, a, 1)
            pred = m_lin * v + c_lin
            ss_res = float(((a - pred) ** 2).sum())
            ss_tot = float(((a - a.mean()) ** 2).sum())
            stats["linear_slope"] = float(m_lin)
            stats["linear_r2"] = float(1 - ss_res / ss_tot) if ss_tot > 0 else np.nan
        except Exception:
            pass

        # Distribution-relative sensitivity: %-area per percentile point of the
        # parameter's own distribution (comparable across sites).
        try:
            pct = v.rank(pct=True) * 100.0
            stats["pctile_slope"] = float(np.polyfit(pct, a, 1)[0])
        except Exception:
            pass

        # Log-linearized relationships require strictly positive values
        pos = mask & (values > 0) & (roi > 0)
        if pos.sum() >= 3:
            lv, la = np.log(values[pos]), np.log(roi[pos])
            stats["log_pearson"] = float(roi[pos].corr(np.log(values[pos]), method="pearson"))
            if lv.nunique() >= 3 and la.nunique() >= 3:
                slope, intercept = np.polyfit(lv, la, 1)
                pred = slope * lv + intercept
                ss_res = float(((la - pred) ** 2).sum())
                ss_tot = float(((la - la.mean()) ** 2).sum())
                stats["loglog_slope"] = float(slope)
                stats["loglog_r2"] = float(1 - ss_res / ss_tot) if ss_tot > 0 else np.nan

        # Rising vs falling limb (hysteresis indicator)
        rising = values.diff() > 0
        for name, limb in (("pearson_rising", rising), ("pearson_falling", ~rising)):
            sel = mask & limb
            if sel.sum() >= 10:
                stats[name] = float(roi[sel].corr(values[sel], method="pearson"))

        # Low vs high half of the parameter range (saturation indicator)
        median = v.median()
        for name, sel in (("pearson_low_half", mask & (values <= median)),
                          ("pearson_high_half", mask & (values > median))):
            if sel.sum() >= 10:
                stats[name] = float(roi[sel].corr(values[sel], method="pearson"))

        stats["assessment"] = self._narrative(stats)
        return stats

    # ==================================================================================================================
    #
    # ==================================================================================================================
    @staticmethod
    def _narrative(s: dict) -> str:
        """Rule-based interpretation of the statistics for one parameter.
        Heuristic by design: describes what the numbers are consistent with."""
        code = s["code"]
        kind = CODE_HINTS.get(code, "this parameter")
        r, rho = s["pearson"], s["spearman"]
        parts = []

        def f(x):
            return "n/a" if x is None or (isinstance(x, float) and np.isnan(x)) else f"{x:.2f}"

        # Strength/shape of the relationship
        if not np.isnan(r) and not np.isnan(rho):
            if abs(r) >= 0.7:
                parts.append(f"ROI water area shows a strong, approximately linear "
                             f"relationship with {kind} (Pearson r={f(r)}, Spearman rho={f(rho)}).")
            elif abs(rho) >= 0.7 > abs(r):
                parts.append(f"ROI water area is strongly but nonlinearly related to {kind} "
                             f"(Spearman rho={f(rho)} vs Pearson r={f(r)}): the relationship is "
                             f"monotonic but curved, so linear correlation understates it.")
                if not np.isnan(s["log_pearson"]) and abs(s["log_pearson"]) > abs(r) + 0.1:
                    parts.append(f"Correlating against log({kind}) recovers much of the "
                                 f"relationship (r={f(s['log_pearson'])}), consistent with a "
                                 f"power-law (rating-curve-like) link.")
            elif max(abs(r), abs(rho)) >= 0.4:
                parts.append(f"ROI water area is moderately related to {kind} "
                             f"(Pearson r={f(r)}, Spearman rho={f(rho)}).")
            else:
                parts.append(f"ROI water area shows little direct relationship with {kind} "
                             f"(Pearson r={f(r)}, Spearman rho={f(rho)}).")

        # Power-law fit
        if not np.isnan(s["loglog_slope"]) and not np.isnan(s["loglog_r2"]) and s["loglog_r2"] >= 0.5:
            parts.append(f"A power-law fit area ~ {kind}^b gives b={s['loglog_slope']:.2f} "
                         f"(log-log R^2={s['loglog_r2']:.2f}).")

        # Saturation
        lo, hi = s["pearson_low_half"], s["pearson_high_half"]
        if not np.isnan(lo) and not np.isnan(hi) and abs(lo) - abs(hi) >= 0.25 and abs(lo) >= 0.5:
            parts.append(f"The relationship is much stronger in the lower half of the {kind} "
                         f"range (r={f(lo)}) than the upper half (r={f(hi)}), consistent with "
                         f"the ROI's wetted area saturating at higher values - e.g. flow "
                         f"confined to banks or the ROI already fully wetted.")

        # Hysteresis - narrated for discharge only per the agreed methodology
        # (loop rating on mobile beds); numbers stay in the table for all.
        ri, fa = s["pearson_rising"], s["pearson_falling"]
        if code == "00060" and not np.isnan(ri) and not np.isnan(fa) and abs(ri - fa) >= 0.2:
            parts.append(f"Rising-limb (r={f(ri)}) and falling-limb (r={f(fa)}) correlations "
                         f"differ noticeably, suggesting hysteresis: the same {kind} value "
                         f"corresponds to different water extents on rising vs falling stages.")

        # Parameter-specific context
        if code == "00065":
            if not np.isnan(s["linear_slope"]):
                form = "linear"
                if (not np.isnan(s["loglog_r2"]) and not np.isnan(s["linear_r2"])
                        and s["loglog_r2"] > s["linear_r2"] + 0.05):
                    form = f"power law (b={s['loglog_slope']:.2f})"
                sens = f"Sensitivity: {s['linear_slope']:.2f} %-area per unit of stage"
                if not np.isnan(s["pctile_slope"]):
                    sens += f", {s['pctile_slope']:.3f} %-area per stage percentile"
                sens += (f". Better-fitting form: {form} "
                         f"(linear R^2={s['linear_r2']:.2f}"
                         + ("" if np.isnan(s["loglog_r2"]) else f", log-log R^2={s['loglog_r2']:.2f}") + ").")
                parts.append(sens)
            parts.append("Gage height directly controls where the water surface intersects the "
                         "channel geometry, so a strong monotonic link with segmented area is "
                         "the physically expected result.")
        elif code == "00060":
            parts.append("Raw Pearson understates this relationship by construction; rank "
                         "statistics (Spearman) are authoritative for discharge.")
            parts.append("Discharge at most gages is computed from stage via a nonlinear, "
                         "periodically shifted rating curve (and may exhibit loop-rating "
                         "hysteresis on mobile beds), so pixel area typically tracks stage far "
                         "better than raw discharge; compare against log(discharge) or the "
                         "power-law fit above rather than the linear coefficient.")
        elif code in ("00010", "00020"):
            parts.append("Temperature has no direct hydraulic control on wetted area; any "
                         "correlation observed here more likely reflects shared seasonal or "
                         "diurnal trends than a causal link.")

        return " ".join(parts) if parts else "No assessment available."

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _write_workbook(self, xlsx_path, work, roi_col, param_cols, all_stats) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.chart import ScatterChart, Reference, Series

        wb = Workbook()
        used_titles = set()

        # ---------------- Assessment tab (first) ----------------
        ws = wb.active
        ws.title = "Assessment"
        stat_cols = ["Parameter", "Method", "N pairs", "Pearson r", "Spearman rho",
                     "Linear slope (%/unit)", "Linear R^2", "Slope (%/percentile)",
                     "r vs log(value)", "Power-law b", "log-log R^2",
                     "r rising limb", "r falling limb",
                     "r low half", "r high half"]
        ws.append(stat_cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for s in all_stats:
            ws.append([
                s["column"], METHODS.get(s["code"], METHOD_DEFAULT), s["n"],
                *[None if np.isnan(x) else round(x, 3) for x in
                  (s["pearson"], s["spearman"],
                   s["linear_slope"], s["linear_r2"], s["pctile_slope"],
                   s["log_pearson"],
                   s["loglog_slope"], s["loglog_r2"],
                   s["pearson_rising"], s["pearson_falling"],
                   s["pearson_low_half"], s["pearson_high_half"])],
            ])
        ws.append([])
        ws.append(["Site assessment (heuristic - describes what the statistics are "
                   "consistent with, not established causation):"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for s in all_stats:
            ws.append([s["column"]])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.append([s["assessment"]])
            ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[ws.max_row].height = 60
            ws.append([])
        ws.column_dimensions["A"].width = 110
        ws.column_dimensions["B"].width = 42
        for col_letter in "CDEFGHIJKLMNO":
            ws.column_dimensions[col_letter].width = 13
        ws.freeze_panes = "A2"

        # ---------------- One tab per parameter ----------------
        import os as _os
        filenames = (work["Image Path"].astype(str).map(_os.path.basename)
                     if "Image Path" in work.columns else pd.Series([""] * len(work)))

        for col in param_cols:
            title = self._sheet_title(col, used_titles)
            pws = wb.create_sheet(title)
            pws.append(["Index", "Timestamp", roi_col, col, "Image Filename"])
            for cell in pws[1]:
                cell.font = Font(bold=True)

            values = pd.to_numeric(work[col], errors="coerce")
            for i in range(len(work)):
                v = values.iloc[i]
                a = work["_roi"].iloc[i]
                ts = work["_ts"].iloc[i]
                pws.append([
                    i + 1,   # 1-based: matches the filmstrip index labels
                    ts.strftime("%Y-%m-%d %H:%M:%S") if isinstance(ts, pd.Timestamp) and pd.notna(ts) else "",
                    None if pd.isna(a) else float(a),
                    None if pd.isna(v) else float(v),
                    filenames.iloc[i],
                ])

            n_rows = len(work) + 1

            # --- Chart 1: time-series overlay ---
            # Per-parameter treatment: stage (00065) and ROI area are both
            # single-unit scalars of comparable magnitude, so they plot RAW on
            # one shared axis. Discharge (a rate) and other parameters span
            # incompatible ranges, so each series is normalized to its own 0-1
            # range on one shared axis. Raw values always remain in C and D.
            from openpyxl.chart import LineChart
            m_code = PARAM_COL_RE.search(str(col))
            overlay_code = m_code.group(2) if m_code else ""
            values_norm_src = pd.to_numeric(work[col], errors="coerce")
            a_src = work["_roi"]

            if overlay_code == "00065":
                ts_chart = LineChart()
                ts_chart.title = f"{roi_col} and {col} over time"
                ts_chart.style = 12
                ts_chart.height, ts_chart.width = 12, 24
                ts_chart.y_axis.title = "Value  (ROI area in %, gage height in ft)"
                ts_chart.x_axis.title = "Image index"
                ts_chart.x_axis.delete = False
                ts_chart.y_axis.delete = False
                ts_chart.y_axis.scaling.min = 0
                data_ref = Reference(pws, min_col=3, max_col=4, min_row=1, max_row=n_rows)
                ts_chart.add_data(data_ref, titles_from_data=True)
                cats = Reference(pws, min_col=1, min_row=2, max_row=n_rows)
                ts_chart.set_categories(cats)
                for ser in ts_chart.series:
                    ser.marker.symbol = "none"
                    ser.smooth = False
                pws.add_chart(ts_chart, "F2")
            elif overlay_code == "00060":
                # Discharge - two views (agreed methodology):
                # (a) Percentile traces: both series as percentiles of their
                #     own distribution - the direct picture of rank tracking.
                #     Percentile columns G/H are written by the scatter block.
                pct_chart = LineChart()
                pct_chart.title = f"{roi_col} and {col} over time - PERCENTILES"
                pct_chart.style = 12
                pct_chart.height, pct_chart.width = 12, 24
                pct_chart.y_axis.title = "Percentile of each series' own distribution"
                pct_chart.x_axis.title = "Image index"
                pct_chart.x_axis.delete = False
                pct_chart.y_axis.delete = False
                pct_chart.y_axis.scaling.min = 0
                pct_chart.y_axis.scaling.max = 100
                pref = Reference(pws, min_col=7, max_col=8, min_row=1, max_row=n_rows)
                pct_chart.add_data(pref, titles_from_data=True)
                cats = Reference(pws, min_col=1, min_row=2, max_row=n_rows)
                pct_chart.set_categories(cats)
                for ser in pct_chart.series:
                    ser.marker.symbol = "none"
                    ser.smooth = False
                pws.add_chart(pct_chart, "F2")

                # (b) Raw ROI area with log10(discharge), one shared axis.
                pws.cell(row=1, column=12, value=f"log10({col})")
                pws.cell(row=1, column=12).font = Font(bold=True)
                for i in range(len(work)):
                    v = values_norm_src.iloc[i]
                    pws.cell(row=i + 2, column=12,
                             value=None if (pd.isna(v) or v <= 0) else round(float(np.log10(v)), 5))
                log_chart = LineChart()
                log_chart.title = f"{roi_col} (raw) and log10({col}) over time"
                log_chart.style = 12
                log_chart.height, log_chart.width = 12, 24
                log_chart.y_axis.title = "Value  (ROI area in %, discharge as log10 cfs)"
                log_chart.x_axis.title = "Image index"
                log_chart.x_axis.delete = False
                log_chart.y_axis.delete = False
                log_chart.y_axis.scaling.min = 0
                aref = Reference(pws, min_col=3, min_row=1, max_row=n_rows)
                log_chart.add_data(aref, titles_from_data=True)
                lref = Reference(pws, min_col=12, min_row=1, max_row=n_rows)
                log_chart.add_data(lref, titles_from_data=True)
                cats2 = Reference(pws, min_col=1, min_row=2, max_row=n_rows)
                log_chart.set_categories(cats2)
                for ser in log_chart.series:
                    ser.marker.symbol = "none"
                    ser.smooth = False
                pws.add_chart(log_chart, "F28")

            else:

                def _norm(series):
                    lo, hi = series.min(), series.max()
                    rng = hi - lo
                    return (series - lo) / rng if rng and not pd.isna(rng) else series * 0.0

                a_n, v_n = _norm(a_src), _norm(values_norm_src)
                pws.cell(row=1, column=10, value=f"{roi_col} (normalized 0-1)")
                pws.cell(row=1, column=11, value=f"{col} (normalized 0-1)")
                pws.cell(row=1, column=10).font = Font(bold=True)
                pws.cell(row=1, column=11).font = Font(bold=True)
                for i in range(len(work)):
                    pws.cell(row=i + 2, column=10,
                             value=None if pd.isna(a_n.iloc[i]) else round(float(a_n.iloc[i]), 5))
                    pws.cell(row=i + 2, column=11,
                             value=None if pd.isna(v_n.iloc[i]) else round(float(v_n.iloc[i]), 5))

                ts_chart = LineChart()
                ts_chart.title = f"{roi_col} and {col} over time (each normalized to its own 0-1 range)"
                ts_chart.style = 12
                ts_chart.height, ts_chart.width = 12, 24
                ts_chart.y_axis.title = "Normalized value (0 = series min, 1 = series max)"
                ts_chart.x_axis.title = "Image index"
                ts_chart.x_axis.delete = False
                ts_chart.y_axis.delete = False
                ts_chart.y_axis.scaling.min = 0
                ts_chart.y_axis.scaling.max = 1
                data_ref = Reference(pws, min_col=10, max_col=11, min_row=1, max_row=n_rows)
                ts_chart.add_data(data_ref, titles_from_data=True)
                cats = Reference(pws, min_col=1, min_row=2, max_row=n_rows)  # Index as x categories
                ts_chart.set_categories(cats)
                for ser in ts_chart.series:
                    ser.marker.symbol = "none"
                    ser.smooth = False
                pws.add_chart(ts_chart, "F2")

            # --- Chart 2: per-methodology correlation chart ---
            m = PARAM_COL_RE.search(str(col))
            code = m.group(2) if m else ""
            values_num = pd.to_numeric(work[col], errors="coerce")

            if code == "00060":
                # Discharge: rank-rank (percentile) scatter - rank statistics
                # are authoritative for this parameter.
                pws.cell(row=1, column=7, value="ROI Area Percentile")
                pws.cell(row=1, column=8, value=f"{col} Percentile")
                pws.cell(row=1, column=7).font = Font(bold=True)
                pws.cell(row=1, column=8).font = Font(bold=True)
                a_pct = work["_roi"].rank(pct=True) * 100.0
                v_pct = values_num.rank(pct=True) * 100.0
                for i in range(len(work)):
                    pws.cell(row=i + 2, column=7,
                             value=None if pd.isna(a_pct.iloc[i]) else float(a_pct.iloc[i]))
                    pws.cell(row=i + 2, column=8,
                             value=None if pd.isna(v_pct.iloc[i]) else float(v_pct.iloc[i]))
                chart = ScatterChart()
                chart.title = f"{col} vs {roi_col} - PERCENTILES (rank-based view)"
                chart.x_axis.title = f"{roi_col} percentile"
                chart.y_axis.title = f"{col} percentile"
                chart.x_axis.delete = False
                chart.y_axis.delete = False
                xref = Reference(pws, min_col=7, min_row=2, max_row=n_rows)
                yref = Reference(pws, min_col=8, min_row=2, max_row=n_rows)
            else:
                # Stage and everything else: raw-value scatter. For stage the
                # fitted curve (chosen form by R^2) is drawn on the chart.
                chart = ScatterChart()
                chart.title = f"{col} vs {roi_col}"
                chart.x_axis.title = roi_col
                chart.y_axis.title = col
                chart.x_axis.delete = False
                chart.y_axis.delete = False
                xref = Reference(pws, min_col=3, min_row=2, max_row=n_rows)
                yref = Reference(pws, min_col=4, min_row=2, max_row=n_rows)

            chart.style = 13
            chart.height, chart.width = 12, 20
            series = Series(yref, xref, title=col)
            series.marker.symbol = "circle"
            series.marker.size = 4
            series.graphicalProperties.line.noFill = True   # points, not lines
            chart.series.append(series)

            if code == "00065":
                # Fitted curve for stage: 50-point grid, form chosen by R^2.
                st = next((x for x in all_stats if x["column"] == col), None)
                mask2 = values_num.notna() & work["_roi"].notna()
                vv, aa = values_num[mask2], work["_roi"][mask2]
                if st is not None and len(vv) >= 3 and not np.isnan(st.get("linear_slope", np.nan)):
                    use_power = (not np.isnan(st["loglog_r2"]) and not np.isnan(st["linear_r2"])
                                 and st["loglog_r2"] > st["linear_r2"] + 0.05 and (vv > 0).all())
                    grid = np.linspace(float(vv.min()), float(vv.max()), 50)
                    if use_power:
                        pos = (vv > 0) & (aa > 0)
                        d, c0 = np.polyfit(np.log(vv[pos]), np.log(aa[pos]), 1)
                        fit_area = np.exp(c0) * np.power(np.clip(grid, 1e-9, None), d)
                        fit_label = "Power-law fit"
                    else:
                        m1, c1 = np.polyfit(vv, aa, 1)
                        fit_area = m1 * grid + c1
                        fit_label = "Linear fit"
                    pws.cell(row=1, column=7, value=f"{fit_label}: fitted ROI area (%)")
                    pws.cell(row=1, column=8, value=f"{fit_label}: stage grid")
                    pws.cell(row=1, column=7).font = Font(bold=True)
                    pws.cell(row=1, column=8).font = Font(bold=True)
                    for i in range(len(grid)):
                        pws.cell(row=i + 2, column=7, value=float(fit_area[i]))
                        pws.cell(row=i + 2, column=8, value=float(grid[i]))
                    fx = Reference(pws, min_col=7, min_row=2, max_row=len(grid) + 1)
                    fy = Reference(pws, min_col=8, min_row=2, max_row=len(grid) + 1)
                    fit_series = Series(fy, fx, title=fit_label)
                    fit_series.marker.symbol = "none"
                    chart.series.append(fit_series)

            chart.varyColors = False   # uniform points - not one color/legend entry per point
            chart.legend = None if code != "00065" else chart.legend
            pws.add_chart(chart, "F54" if code == "00060" else "F28")

            pws.column_dimensions["A"].width = 8
            pws.column_dimensions["B"].width = 20
            pws.column_dimensions["C"].width = 18
            pws.column_dimensions["D"].width = 18
            pws.column_dimensions["E"].width = 55
            pws.freeze_panes = "A2"

        wb.save(xlsx_path)

    # ==================================================================================================================
    #
    # ==================================================================================================================
    @staticmethod
    def _sheet_title(col: str, used: set) -> str:
        """Excel-legal sheet title: strip []:*?/\\ and fit 31 chars, unique."""
        m = PARAM_COL_RE.search(str(col))
        code = m.group(2) if m else ""
        base = PARAM_COL_RE.sub("", str(col)).strip()
        base = re.sub(r"[\[\]:*?/\\]", "", base)
        title = f"{base} {code}".strip()
        if len(title) > 31:
            title = f"{base[:31 - len(code) - 2].rstrip(', ')} {code}"[:31]
        candidate, k = title, 2
        while candidate.lower() in used:
            suffix = f" ({k})"
            candidate = title[:31 - len(suffix)] + suffix
            k += 1
        used.add(candidate.lower())
        return candidate
