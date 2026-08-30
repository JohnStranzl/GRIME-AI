#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Author: John Edward Stranzl, Jr.
# Affiliation(s): University of Nebraska-Lincoln, Blade Vision Systems, LLC
# Contact: jstranzl2@huskers.unl.edu, johnstranzl@gmail.com
# License: Apache License, Version 2.0, http://www.apache.org/licenses/LICENSE-2.0

"""
Correlate USGS HIVIS camera images with co-located NWIS sensor data.

Images:  filenames of the form  SITE_NAME___YYYY-MM-DDTHH-MM-SSZ.jpg  (UTC)
Sensors: either the raw NWIS RDB .txt download or the reformatted .csv,
         with columns agency_cd, site_no, datetime, tz_cd, then one value
         column and one _cd qualifier column per parameter. Parameter values
         may be sparse - a sensor can go offline while others keep reporting.

Both are converted to UTC and matched with a configurable tolerance
(default: half the detected sensor sampling interval). The report flags
  - images with no sensor reading within tolerance,
  - sensor readings with no image within tolerance,
  - gaps in each stream where data is missing entirely.

Outputs (same stem, written to the image folder unless overridden):
  SensorImageCorrelation_<ts>.csv    image-correlation table
  SensorImageCorrelation_<ts>.xlsx   Image Correlation / Sensor Coverage /
                                     Gaps / Summary worksheets
"""

import os
import re
import json
from datetime import datetime, timedelta, timezone
from typing import List, Optional, Tuple

import pandas as pd

# Fixed UTC offsets for the timezone codes NWIS emits in tz_cd.
TZ_OFFSETS_HOURS = {
    "UTC": 0, "GMT": 0,
    "AST": -4, "ADT": -3,
    "EST": -5, "EDT": -4,
    "CST": -6, "CDT": -5,
    "MST": -7, "MDT": -6,
    "PST": -8, "PDT": -7,
    "AKST": -9, "AKDT": -8,
    "HST": -10, "HDT": -9,
}

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png")
IMAGE_TS_RE = re.compile(r"___(\d{4}-\d{2}-\d{2}T\d{2}-\d{2}-\d{2})Z", re.IGNORECASE)


class GRIME_AI_SensorImageCorrelator:

    def __init__(self, show_gui: bool = False):
        self.className = "GRIME_AI_SensorImageCorrelator"
        self.show_gui  = show_gui

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def correlate(self,
                  image_folder: str,
                  sensor_csv: str,
                  output_folder: Optional[str] = None,
                  tolerance_minutes: Optional[float] = None,
                  recursive: bool = False,
                  progress=None) -> Tuple[str, str]:
        """
        Correlate image timestamps with sensor timestamps and write the report.

        Parameters
        ----------
        image_folder : str
            Folder containing downloaded HIVIS images.
        sensor_csv : str
            Path to the sidecar sensor CSV (reformatted NWIS RDB file).
        output_folder : str, optional
            Where to write the report files. Defaults to image_folder.
        tolerance_minutes : float, optional
            Maximum |image time - sensor time| for a match. Defaults to half
            the detected sensor sampling interval.
        recursive : bool
            Recurse into subfolders of image_folder.
        progress : callable, optional
            progress(done, total, label) - called periodically so a GUI can
            show a progress bar. Never required; ignored when None.

        Returns
        -------
        (csv_path, xlsx_path)
        """
        def report(done, total, label):
            if progress:
                try:
                    progress(done, total, label)
                except Exception:
                    pass

        report(0, 1, "Scanning images...")
        images  = self._scan_images(image_folder, recursive)
        report(0, 1, "Reading sensor data...")
        sensors = self._read_sensor_file(sensor_csv)

        if images.empty and sensors.empty:
            raise ValueError("No parseable images found and no sensor rows found - nothing to correlate.")

        sensor_interval = self._detect_interval(sensors["utc"]) if not sensors.empty else None
        image_interval  = self._detect_interval(images["utc"])  if not images.empty  else None

        if tolerance_minutes is None:
            tolerance = (sensor_interval / 2) if sensor_interval is not None else timedelta(minutes=7.5)
        else:
            tolerance = timedelta(minutes=float(tolerance_minutes))

        total = max(len(images) + len(sensors), 1)
        image_report  = self._match_images_to_sensors(
            images, sensors, tolerance,
            progress=lambda i: report(i, total, "Matching images to sensor data..."))
        sensor_report = self._match_sensors_to_images(
            sensors, images, tolerance,
            progress=lambda i: report(len(images) + i, total, "Matching sensor data to images..."))
        report(total, total, "Writing report...")
        gaps          = self._find_gaps(images, sensors, image_interval, sensor_interval)
        summary       = self._build_summary(image_folder, sensor_csv, images, sensors,
                                            image_interval, sensor_interval, tolerance,
                                            image_report, sensor_report, gaps)

        out_dir = output_folder or image_folder
        os.makedirs(out_dir, exist_ok=True)
        stem      = "SensorImageCorrelation_" + datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path  = os.path.join(out_dir, stem + ".csv")
        xlsx_path = os.path.join(out_dir, stem + ".xlsx")

        image_report.to_csv(csv_path, index=False)
        self._write_xlsx(xlsx_path, image_report, sensor_report, gaps, summary,
                         progress=lambda label: report(total, total, label))
        report(total, total, "Done")

        return csv_path, xlsx_path

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def sensor_values_for_images(self, image_paths: List[str], sensor_file: str,
                                 tolerance_minutes: Optional[float] = None) -> pd.DataFrame:
        """Return sensor values aligned to a list of image paths, one row per
        path in the same order. For use by feature-extraction pipelines that
        want sensor columns appended to per-image rows.

        Columns: Sensor Time (UTC), Sensor Offset (s), Sensor Match, then one
        column per parameter (named from the RDB header descriptions when
        available). Sensor Match is Matched / Matched (partial) /
        Matched (no values) / No sensor data / No timestamp.
        """
        sensors  = self._read_sensor_file(sensor_file)
        interval = self._detect_interval(sensors["utc"]) if not sensors.empty else None
        if tolerance_minutes is None:
            tolerance = (interval / 2) if interval is not None else timedelta(minutes=7.5)
        else:
            tolerance = timedelta(minutes=float(tolerance_minutes))
        sensor_times = sensors["utc"] if not sensors.empty else pd.Series(dtype="object")

        out_rows = []
        for path in image_paths:
            ts = None
            m = IMAGE_TS_RE.search(os.path.basename(str(path)))
            if m:
                try:
                    ts = datetime.strptime(m.group(1), "%Y-%m-%dT%H-%M-%S").replace(tzinfo=timezone.utc)
                except ValueError:
                    ts = None

            if ts is None:
                row = {"Sensor Time (UTC)": "", "Sensor Offset (s)": "",
                       "Sensor Match": "No timestamp"}
                row.update({c: "" for c in self._param_columns})
            else:
                idx, offset = self._nearest(ts, sensor_times)
                matched = idx is not None and abs(offset) <= tolerance
                if matched:
                    values  = {c: sensors.loc[idx, c] for c in self._param_columns}
                    present = sum(1 for v in values.values() if self._has_value(v))
                    if present == len(self._param_columns):
                        status = "Matched"
                    elif present:
                        status = "Matched (partial)"
                    else:
                        status = "Matched (no values)"
                else:
                    values = {c: "" for c in self._param_columns}
                    status = "No sensor data"
                row = {
                    "Sensor Time (UTC)": sensors.loc[idx, "utc"].strftime("%Y-%m-%d %H:%M:%S") if idx is not None else "",
                    "Sensor Offset (s)": round(offset.total_seconds(), 1) if offset is not None else "",
                    "Sensor Match":      status,
                }
                row.update(values)
            out_rows.append(row)

        cols = ["Sensor Time (UTC)", "Sensor Offset (s)", "Sensor Match"] + self._param_columns
        return pd.DataFrame(out_rows, columns=cols)

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _scan_images(self, folder: str, recursive: bool) -> pd.DataFrame:
        """Collect image files and parse the UTC timestamp from each filename."""
        rows, unparseable = [], []

        if recursive:
            walker = ((os.path.join(dirpath, f) for f in files)
                      for dirpath, _dirs, files in os.walk(folder))
            paths = [p for gen in walker for p in gen]
        else:
            paths = [os.path.join(folder, f) for f in os.listdir(folder)] if os.path.isdir(folder) else []

        for path in sorted(paths):
            name = os.path.basename(path)
            if os.path.splitext(name)[1].lower() not in IMAGE_EXTENSIONS:
                continue
            m = IMAGE_TS_RE.search(name)
            if not m:
                unparseable.append(name)
                continue
            try:
                ts = datetime.strptime(m.group(1), "%Y-%m-%dT%H-%M-%S").replace(tzinfo=timezone.utc)
            except ValueError:
                unparseable.append(name)
                continue
            rows.append({"filename": name, "path": path, "utc": ts})

        if unparseable:
            print(f"[{self.className}] Warning: {len(unparseable)} image filenames without a "
                  f"parseable ___YYYY-MM-DDTHH-MM-SSZ timestamp were skipped.")

        df = pd.DataFrame(rows, columns=["filename", "path", "utc"])
        self._unparseable_images = unparseable
        return df.sort_values("utc").reset_index(drop=True) if not df.empty else df

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _read_sensor_file(self, sensor_file: str) -> pd.DataFrame:
        """Read the sensor sidecar file (raw NWIS RDB .txt or reformatted .csv)
        and convert timestamps to UTC."""
        if sensor_file.lower().endswith(".txt"):
            # Raw RDB: tab-separated, '#' comment header
            df = pd.read_csv(sensor_file, delimiter="\t", comment="#", dtype=str)
            self._param_labels = self._parse_rdb_parameter_descriptions(sensor_file)
        else:
            df = pd.read_csv(sensor_file, dtype=str)
            # The reformatted CSV has no header comments; if the original RDB
            # .txt sits alongside it, harvest the parameter descriptions there.
            sibling_txt = os.path.splitext(sensor_file)[0] + ".txt"
            self._param_labels = (self._parse_rdb_parameter_descriptions(sibling_txt)
                                  if os.path.isfile(sibling_txt) else {})
        if "datetime" not in df.columns:
            raise ValueError(f"Sensor file {sensor_file} has no 'datetime' column - "
                             "is this an NWIS RDB download or its reformatted CSV?")
        # Drop the RDB column-format row (e.g. '5s 15s 20d ...') if present
        if "agency_cd" in df.columns:
            df = df[~df["agency_cd"].astype(str).str.contains("5s", na=False)]

        has_tz = "tz_cd" in df.columns
        unknown_tz = set()

        def to_utc(row):
            try:
                local = datetime.strptime(str(row["datetime"]).strip(), "%Y-%m-%d %H:%M")
            except ValueError:
                try:
                    local = datetime.strptime(str(row["datetime"]).strip(), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    return pd.NaT
            offset = 0
            if has_tz:
                code = str(row["tz_cd"]).strip().upper()
                if code in TZ_OFFSETS_HOURS:
                    offset = TZ_OFFSETS_HOURS[code]
                else:
                    unknown_tz.add(code)
            return local.replace(tzinfo=timezone(timedelta(hours=offset))).astimezone(timezone.utc)

        df["utc"] = df.apply(to_utc, axis=1)
        bad = df["utc"].isna().sum()
        if bad:
            print(f"[{self.className}] Warning: {bad} sensor rows had unparseable timestamps and were dropped.")
        if unknown_tz:
            print(f"[{self.className}] Warning: unrecognized tz_cd value(s) {sorted(unknown_tz)} treated as UTC.")
        self._sensor_bad_rows = int(bad)
        self._unknown_tz = sorted(unknown_tz)

        df = df.dropna(subset=["utc"]).sort_values("utc").reset_index(drop=True)

        # Parameter (value) columns: everything except bookkeeping and *_cd qualifier columns
        bookkeeping = {"agency_cd", "site_no", "datetime", "tz_cd", "utc"}
        raw_params = [c for c in df.columns
                      if c not in bookkeeping and not c.endswith("_cd")]

        # Use the human-readable parameter descriptions from the RDB header
        # as column names, keeping the raw id for traceability.
        renames = {}
        for col in raw_params:
            desc = self._param_labels.get(col)
            if desc:
                renames[col] = f"{desc} [{col}]"
        if renames:
            df = df.rename(columns=renames)
        self._param_columns = [renames.get(c, c) for c in raw_params]
        return df

    # ==================================================================================================================
    #
    # ==================================================================================================================
    @staticmethod
    def _parse_rdb_parameter_descriptions(rdb_txt: str) -> dict:
        """Extract {ts_id}_{param} -> description from the RDB comment header.

        Parses lines of the form:
            #    94479       00010     Temperature, water, degrees Celsius
        """
        labels = {}
        pat = re.compile(r"^#\s+(\d+)\s+(\d{5})\s+(.+?)\s*$")
        try:
            with open(rdb_txt, "r", encoding="utf-8", errors="replace") as fh:
                for line in fh:
                    if not line.startswith("#"):
                        break  # header comments end where the data begins
                    m = pat.match(line)
                    if m:
                        ts_id, param, desc = m.groups()
                        if desc.upper() != "DESCRIPTION":
                            labels[f"{ts_id}_{param}"] = desc.strip()
        except OSError:
            pass
        return labels

    # ==================================================================================================================
    #
    # ==================================================================================================================
    @staticmethod
    def _detect_interval(times: pd.Series) -> Optional[timedelta]:
        """Median spacing between consecutive timestamps, or None if < 2 samples."""
        if len(times) < 2:
            return None
        diffs = times.diff().dropna()
        med = diffs.median()
        return med.to_pytimedelta() if hasattr(med, "to_pytimedelta") else med

    # ==================================================================================================================
    #
    # ==================================================================================================================
    @staticmethod
    def _has_value(v) -> bool:
        """True when a parameter cell holds an actual reading (not blank/NaN)."""
        return not (pd.isna(v) or str(v).strip() == "")

    # ==================================================================================================================
    #
    # ==================================================================================================================
    @staticmethod
    def _nearest(target: datetime, candidates: pd.Series) -> Tuple[Optional[int], Optional[timedelta]]:
        """Index and signed offset of the candidate nearest to target."""
        if candidates.empty:
            return None, None
        deltas = (candidates - target).abs()
        idx = deltas.idxmin()
        return idx, candidates.loc[idx] - target

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _match_images_to_sensors(self, images: pd.DataFrame, sensors: pd.DataFrame,
                                 tolerance: timedelta, progress=None) -> pd.DataFrame:
        rows = []
        sensor_times = sensors["utc"] if not sensors.empty else pd.Series(dtype="object")
        for n, (_, img) in enumerate(images.iterrows()):
            if progress and n % 25 == 0:
                progress(n)
            idx, offset = self._nearest(img["utc"], sensor_times)
            matched = idx is not None and abs(offset) <= tolerance
            if matched:
                values  = {col: sensors.loc[idx, col] for col in self._param_columns}
                present = sum(1 for v in values.values() if self._has_value(v))
                if present == 0:
                    status = "Matched (no values)"
                elif present < len(self._param_columns):
                    status = "Matched (partial)"
                else:
                    status = "Matched"
            else:
                values = {col: "" for col in self._param_columns}
                status = "No sensor data"
            row = {
                "Image Filename":       img["filename"],
                "Image Time (UTC)":     img["utc"].strftime("%Y-%m-%d %H:%M:%S"),
                "Nearest Sensor (UTC)": sensors.loc[idx, "utc"].strftime("%Y-%m-%d %H:%M:%S") if idx is not None else "",
                "Offset (s)":           round(offset.total_seconds(), 1) if offset is not None else "",
                "Status":               status,
            }
            row.update(values)
            rows.append(row)
        cols = (["Image Filename", "Image Time (UTC)", "Nearest Sensor (UTC)", "Offset (s)", "Status"]
                + self._param_columns)
        return pd.DataFrame(rows, columns=cols)

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _match_sensors_to_images(self, sensors: pd.DataFrame, images: pd.DataFrame,
                                 tolerance: timedelta, progress=None) -> pd.DataFrame:
        rows = []
        image_times = images["utc"] if not images.empty else pd.Series(dtype="object")
        for n, (_, sen) in enumerate(sensors.iterrows()):
            if progress and n % 200 == 0:
                progress(n)
            idx, offset = self._nearest(sen["utc"], image_times)
            matched = idx is not None and abs(offset) <= tolerance
            row = {
                "Sensor Time (UTC)":   sen["utc"].strftime("%Y-%m-%d %H:%M:%S"),
                "Sensor Time (local)": f'{sen.get("datetime", "")} {sen.get("tz_cd", "")}'.strip(),
                "Nearest Image":       images.loc[idx, "filename"] if idx is not None else "",
                "Offset (s)":          round(offset.total_seconds(), 1) if offset is not None else "",
                "Status":              "Matched" if matched else "No image",
            }
            for col in self._param_columns:
                row[col] = sen[col]
            rows.append(row)
        cols = (["Sensor Time (UTC)", "Sensor Time (local)", "Nearest Image", "Offset (s)", "Status"]
                + self._param_columns)
        return pd.DataFrame(rows, columns=cols)

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _find_gaps(self, images: pd.DataFrame, sensors: pd.DataFrame,
                   image_interval: Optional[timedelta],
                   sensor_interval: Optional[timedelta]) -> pd.DataFrame:
        """Holes in each stream larger than 1.5x its own median interval."""
        rows = []

        def scan(times: pd.Series, interval: Optional[timedelta], label: str):
            if interval is None or len(times) < 2:
                return
            threshold = interval * 1.5
            prev = times.iloc[0]
            for t in times.iloc[1:]:
                delta = t - prev
                if delta > threshold:
                    missed = max(int(round(delta / interval)) - 1, 1)
                    rows.append({
                        "Stream":          label,
                        "Gap Start (UTC)": prev.strftime("%Y-%m-%d %H:%M:%S"),
                        "Gap End (UTC)":   t.strftime("%Y-%m-%d %H:%M:%S"),
                        "Duration (min)":  round(delta.total_seconds() / 60.0, 1),
                        "Est. Missing Samples": missed,
                    })
                prev = t

        scan(sensors["utc"] if not sensors.empty else pd.Series(dtype="object"), sensor_interval, "Sensor data")
        scan(images["utc"]  if not images.empty  else pd.Series(dtype="object"), image_interval,  "Images")

        return pd.DataFrame(rows, columns=["Stream", "Gap Start (UTC)", "Gap End (UTC)",
                                           "Duration (min)", "Est. Missing Samples"])

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _build_summary(self, image_folder, sensor_csv, images, sensors,
                       image_interval, sensor_interval, tolerance,
                       image_report, sensor_report, gaps) -> dict:
        fmt = lambda td: (round(td.total_seconds() / 60.0, 1) if td is not None else "n/a")
        n_img_full    = int((image_report["Status"] == "Matched").sum()) if not image_report.empty else 0
        n_img_partial = int((image_report["Status"] == "Matched (partial)").sum()) if not image_report.empty else 0
        n_img_novals  = int((image_report["Status"] == "Matched (no values)").sum()) if not image_report.empty else 0
        n_img_matched = n_img_full + n_img_partial + n_img_novals
        n_sen_matched = int((sensor_report["Status"] == "Matched").sum()) if not sensor_report.empty else 0
        return {
            "run_timestamp":                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "image_folder":                 image_folder,
            "sensor_file":                  sensor_csv,
            "images_found":                 int(len(images)),
            "image_filenames_unparseable":  len(getattr(self, "_unparseable_images", [])),
            "sensor_rows":                  int(len(sensors)),
            "sensor_rows_unparseable":      getattr(self, "_sensor_bad_rows", 0),
            **{f"values_present [{c}]": int(sensors[c].map(self._has_value).sum()) if not sensors.empty else 0
               for c in self._param_columns},
            "unrecognized_tz_codes":        ", ".join(getattr(self, "_unknown_tz", [])) or "none",
            "sensor_interval_min":          fmt(sensor_interval),
            "image_interval_min":           fmt(image_interval),
            "match_tolerance_min":          round(tolerance.total_seconds() / 60.0, 2),
            "images_matched":               n_img_matched,
            "images_matched_all_params":    n_img_full,
            "images_matched_partial_params": n_img_partial,
            "images_matched_no_values":     n_img_novals,
            "images_without_sensor_data":   int(len(images)) - n_img_matched,
            "sensor_readings_matched":      n_sen_matched,
            "sensor_readings_without_image": int(len(sensors)) - n_sen_matched,
            "gaps_detected":                int(len(gaps)),
        }

    # kept for backward compatibility with earlier revisions of this module
    def _read_sensor_csv(self, sensor_csv):
        return self._read_sensor_file(sensor_csv)

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def _write_xlsx(self, xlsx_path, image_report, sensor_report, gaps, summary,
                    progress=None) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows

        flag_fill    = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # unmatched
        partial_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # partial/empty values
        wb = Workbook()

        def pump(label):
            if progress:
                try:
                    progress(label)
                except Exception:
                    pass

        def add_sheet(title, df, status_col=None, first=False):
            pump(f"Writing report - {title}...")
            ws = wb.active if first else wb.create_sheet(title)
            ws.title = title
            for n, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
                ws.append(r)
                if n % 1000 == 0:
                    pump(f"Writing report - {title}...")
            for cell in ws[1]:
                cell.font = Font(bold=True)
            if status_col and not df.empty:
                # NOTE: ws.max_row / ws.max_column are O(cells) scans in
                # openpyxl - never read them inside a loop. Bound once and
                # walk the cells directly.
                status_idx = list(df.columns).index(status_col)
                n_rows, n_cols = len(df) + 1, len(df.columns)
                for n, row in enumerate(ws.iter_rows(min_row=2, max_row=n_rows, max_col=n_cols)):
                    if n % 1000 == 0:
                        pump(f"Writing report - highlighting {title}...")
                    status = str(row[status_idx].value or "")
                    if status == "Matched":
                        continue
                    fill = partial_fill if status.startswith("Matched (") else flag_fill
                    for cell in row:
                        cell.fill = fill
            ws.freeze_panes = "A2"
            return ws

        add_sheet("Image Correlation", image_report, status_col="Status", first=True)
        add_sheet("Sensor Coverage",   sensor_report, status_col="Status")
        add_sheet("Gaps",              gaps)

        ws = wb.create_sheet("Summary")
        ws.append(["Item", "Value"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for key, value in summary.items():
            ws.append([key, value])
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 60

        pump("Writing report - saving workbook...")
        wb.save(xlsx_path)


# ======================================================================================================================
# Command-line use:  python GRIME_AI_SensorImageCorrelator.py <image_folder> <sensor_csv> [tolerance_minutes]
# ======================================================================================================================
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    tol = float(sys.argv[3]) if len(sys.argv) > 3 else None
    correlator = GRIME_AI_SensorImageCorrelator()
    csv_out, xlsx_out = correlator.correlate(sys.argv[1], sys.argv[2], tolerance_minutes=tol)
    print(f"Report written:\n  {csv_out}\n  {xlsx_out}")
